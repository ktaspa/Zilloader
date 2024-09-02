import pandas as pd
import csv
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import numpy as np
import re

wb = openpyxl.Workbook()
ws = wb.active

with open('zillow_properties.csv') as f:
    reader = csv.reader(f, delimiter=',')

    for row_index, row in enumerate(reader, start=1):
        for column_index, cell_value in enumerate(row, start=1):
            ws.cell(row=row_index, column=column_index).value=cell_value

wb.save('zillow_final.xlsx')

properties = openpyxl.load_workbook('zillow_final.xlsx')
sheet = properties.active

def average():
    values = [cell.value for cell in sheet['A:A'][1:] if cell.value is not None] 
    average = sum(int(value) for value in values) / len(values)
    return f'{average:.2f}'

print('$' + average())

square_feet=[]
def parse_sqft():
    pattern = r'[^,]+,[^,]+,\s*([\d,]+\s*sqft)'

    for cell in sheet['C'][1:]:
        if cell.value:
            match = re.search(pattern, cell.value)
            if match:
                square_feet = match.group(1)  # Extract the matched square footage part
                print(square_feet)
            

parse_sqft() #iterate through each cell and only return the text after the 2nd comma


def scatter_plot():
    x_values = []
    y_values = []

    #to extract square footage
    pattern = r'[^,]+,[^,]+,\s*([\d,]+)\s*sqft'

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        x_value = row[0]  
        square_feet_str = row[2]  
        
        if x_value is not None and square_feet_str:
            match = re.search(pattern, square_feet_str)
            if match:
                sqft_value = int(match.group(1).replace(',', ''))
                x_values.append(x_value)
                y_values.append(sqft_value)

      
    #numpy arrays
    x = np.array([int(x) / 1000 for x in x_values])
    y = np.array(y_values)

    plt.scatter(x, y)
    plt.xlabel('Property Value (in thousands)')
    plt.ylabel('Square Feet')
    plt.title('Listed Properties in 75082')
    plt.show()


scatter_plot()